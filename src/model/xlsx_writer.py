"""Määrittelee luokan XlsxWriter ja sen käyttämät apufunktiot"""
from datetime import datetime
from os import chdir, remove
from sys import stderr

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def _write_months_in_sheet(sheet):
    """Kirjoittaa kuukausien nimet xlsx-tiedostoon."""
    months = ["Tammikuu", "Helmikuu", "Maaliskuu", "Huhtikuu", "Toukokuu", "Kesäkuu",
            "Heinäkuu", "Elokuu", "Syyskuu", "Lokakuu", "Marraskuu", "Joulukuu"]

    for i, month in enumerate(months):
        month_column = chr(ord("B") + i)
        sheet[month_column + str(5)] = month


def _write_categories_in_sheet(sheet, categories):
    """Kirjoittaa laskentataulukkoon käyttäjän antamat
    tilitapahtumakategoriat."""

    for i, category in enumerate(categories):
        sheet["A" + str(i + 6)] = category


def _write_values_in_sheet(sheet, values: list, past_month):
    """Kirjoittaa annetut arvot annettuun laskentataulukkoon annettua kuukauden
    numeroa vastaavaan sarakkeeseen."""
    month_column = chr(ord("A") + past_month)

    for i, value in enumerate(values):
        sheet[month_column + str(i + 6)] = value


class XlsxWriter:
    """Hoitaa xlsx-tiedostoihin liittyvät toiminnot."""

    def __init__(self, save_dir):
        chdir(save_dir)
        # Ohjelmalle annetaan aina edellisen kuukauden tilitapahtumat.
        self.past_month = datetime.today().month - 1

    def init_workbooks(self):
        """Luo uuden xlsx-tiedoston ja alustaa uuden laskentataulukon."""
        new_workbook = Workbook()
        sheet = new_workbook.active
        sheet.title = "taloushistoria"

        sheet["B1"] = "Taloushistoria"
        title_font = Font(size=16, bold=True)
        sheet["B1"].font = title_font

        _write_months_in_sheet(sheet)

        # Ensimmäisellä ajokerralla luodaan toissakuukauden tiedosto, jotta
        # write_month() voi toimia normaalisti
        new_workbook.save("talousseuranta_autom" + str(self.past_month - 1) + ".xlsx")

    def write_month(self, categories_values: dict):
        """Kirjoittaa laskentataulukkoon edellistä kuukautta vastaavat arvot."""

        try:
            # Avataan kahden kuukauden takainen tiedosto muokattavaksi
            workbook = load_workbook("talousseuranta_autom" + str(self.past_month - 1) + ".xlsx")
        except FileNotFoundError:
            # Jos tiedostoa ei ole, eli ohjelmaa suoritetaan ensimmäisen kerran,
            # luodaan uusi xlsx-tiedosto ja kutsutaan metodia uudelleen.
            self.init_workbooks()
            self.write_month(categories_values)
            return

        sheet = workbook["taloushistoria"]

        _write_categories_in_sheet(sheet, list(categories_values.keys()))
        _write_values_in_sheet(sheet, list(categories_values.values()), self.past_month)

        # Tallennetaan tiedosto uudella nimellä, edellinen jää varmuuskopioksi
        workbook.save("talousseuranta_autom" + str(self.past_month) + ".xlsx")

        # Poistetaan vanha varmuuskopio. Ensimmäisella ajokerralla kolmen
        # kuukauden takaista tiedostoa ei löydy
        try:
            remove("talousseuranta_autom" + str(self.past_month - 2) + ".xlsx")
        except FileNotFoundError as error:
            print("Virhe kirjoitettaessa xlsx-asiakirjaa: ", error, file=stderr)
